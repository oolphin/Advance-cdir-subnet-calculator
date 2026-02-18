# services/export_service.py
"""
Service d'export multi-format avec optimisations
"""
import pandas as pd
from io import BytesIO, StringIO
import json
from datetime import datetime
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class ExportService:
    """Service d'export professionnel"""
    
    @staticmethod
    def to_excel(data: pd.DataFrame, stats: Dict[str, Any], 
                 network_type: str, session_id: str) -> BytesIO:
        """
        Exporte en Excel avec mise en forme professionnelle
        
        Args:
            data: DataFrame des résultats
            stats: Statistiques du réseau
            network_type: 'VLAN' ou 'VXLAN'
            session_id: ID de session
            
        Returns:
            BytesIO du fichier Excel
        """
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Onglet 1: Données principales
            data.to_excel(writer, sheet_name='Découpage', index=False)
            
            # Onglet 2: Résumé
            ExportService._create_summary_sheet(writer, data, network_type)
            
            # Onglet 3: Statistiques
            ExportService._create_stats_sheet(writer, stats, session_id)
            
            # Mise en forme
            workbook = writer.book
            ExportService._format_main_sheet(workbook['Découpage'])
            
            # Formater les autres feuilles seulement si elles existent
            if 'Résumé' in workbook.sheetnames:
                ExportService._format_summary_sheet(workbook['Résumé'])
            
            if 'Statistiques' in workbook.sheetnames:
                ExportService._format_stats_sheet(workbook['Statistiques'])
        
        output.seek(0)
        return output
    
    @staticmethod
    def _create_summary_sheet(writer, data: pd.DataFrame, network_type: str):
        """Crée l'onglet de résumé"""
        id_field = 'VLAN ID' if network_type == 'VLAN' else 'VNI'
        name_field = 'Nom VLAN' if network_type == 'VLAN' else 'Nom VXLAN'
        
        # Créer le résumé si les colonnes existent
        if id_field in data.columns and name_field in data.columns:
            summary = data.groupby([id_field, name_field]).agg({
                'Sous-réseau': 'count',
                'Hosts utilisables': 'sum'
            }).reset_index()
            
            summary.columns = [id_field, name_field, 'Nb sous-réseaux', 'Total hosts']
            
            # Filtrer les non attribués
            summary = summary[summary[id_field] != "Non attribué"]
            
            # Créer la feuille même si le résumé est vide
            if len(summary) > 0:
                summary.to_excel(writer, sheet_name='Résumé', index=False)
            else:
                # Créer une feuille vide avec les en-têtes
                empty_summary = pd.DataFrame(columns=[id_field, name_field, 'Nb sous-réseaux', 'Total hosts'])
                empty_summary.to_excel(writer, sheet_name='Résumé', index=False)
    
    @staticmethod
    def _create_stats_sheet(writer, stats: Dict[str, Any], session_id: str):
        """Crée l'onglet de statistiques"""
        stats_data = {
            'Métrique': [
                'Réseau de base',
                'Total IPs',
                'IPs utilisées',
                'IPs libres',
                'Taux d\'utilisation (%)',
                'Nombre de sous-réseaux',
                'Efficacité (%)',
                'Date de génération',
                'ID de session'
            ],
            'Valeur': [
                stats.get('base_network', 'N/A'),
                stats.get('total_ips', 0),
                stats.get('used_ips', 0),
                stats.get('free_ips', 0),
                f"{stats.get('utilization', 0):.2f}",
                stats.get('subnet_count', 0),
                f"{stats.get('efficiency', 0):.2f}",
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                session_id
            ]
        }
        
        stats_df = pd.DataFrame(stats_data)
        stats_df.to_excel(writer, sheet_name='Statistiques', index=False)
    
    @staticmethod
    def _format_main_sheet(worksheet):
        """Mise en forme de l'onglet principal"""
        # En-têtes
        header_fill = PatternFill(start_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Lignes alternées
        alt_fill = PatternFill(start_color="F2F2F2", fill_type="solid")
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            if row_idx % 2 == 0:
                for cell in row:
                    cell.fill = alt_fill
        
        # Ajuster les largeurs
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _format_summary_sheet(worksheet):
        """Mise en forme du résumé"""
        ExportService._format_main_sheet(worksheet)
    
    @staticmethod
    def _format_stats_sheet(worksheet):
        """Mise en forme des statistiques"""
        # En-têtes
        header_fill = PatternFill(start_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Largeurs des colonnes
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 40
        
        # Mettre en gras les métriques
        metric_font = Font(bold=True)
        for row in worksheet.iter_rows(min_row=2, max_col=1):
            row[0].font = metric_font
    
    @staticmethod
    def to_csv(data: pd.DataFrame) -> str:
        """Exporte en CSV"""
        return data.to_csv(index=False)
    
    @staticmethod
    def to_json(data: pd.DataFrame, pretty: bool = True) -> str:
        """Exporte en JSON"""
        records = data.to_dict('records')
        
        if pretty:
            return json.dumps(records, indent=2, default=str, ensure_ascii=False)
        else:
            return json.dumps(records, default=str, ensure_ascii=False)
    
    @staticmethod
    def generate_cisco_config(data: pd.DataFrame, vlans: List[Dict]) -> str:
        """Génère la configuration Cisco IOS"""
        lines = []
        
        # En-tête
        lines.append("!" * 60)
        lines.append(f"! Configuration Cisco IOS")
        lines.append(f"! Généré le {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("!" * 60)
        lines.append("")
        
        # VLANs
        lines.append("! === Configuration des VLANs ===")
        for vlan in vlans:
            lines.append(f"vlan {vlan['vlan_id']}")
            lines.append(f" name {vlan['name']}")
            if vlan.get('description'):
                lines.append(f" description {vlan['description']}")
            lines.append("!")
        
        lines.append("")
        
        # Interfaces SVI
        lines.append("! === Interfaces SVI (Layer 3) ===")
        for _, row in data.iterrows():
            vlan_id = row.get('VLAN ID')
            if vlan_id and vlan_id != "Non attribué":
                lines.append(f"interface Vlan{vlan_id}")
                lines.append(f" description {row.get('Nom sous-réseau', '')}")
                lines.append(f" ip address {row.get('Gateway', '')} {row.get('Masque', '')}")
                lines.append(" no shutdown")
                lines.append("!")
        
        lines.append("")
        lines.append("end")
        lines.append("write memory")
        
        return "\n".join(lines)
    
    @staticmethod
    def generate_linux_vxlan_config(data: pd.DataFrame) -> str:
        """Génère les commandes Linux pour VXLAN"""
        lines = []
        
        lines.append("#!/bin/bash")
        lines.append("#" * 60)
        lines.append("# Configuration VXLAN pour Linux")
        lines.append(f"# Généré le {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("#" * 60)
        lines.append("")
        
        for _, row in data.iterrows():
            vni = row.get('VNI')
            if vni and vni != "Non attribué":
                multicast_ip = row.get('Multicast IP', '239.0.0.1')
                subnet = row.get('Sous-réseau', '')
                gateway = row.get('Gateway', '')
                
                # Extraire le prefix length
                prefix_len = subnet.split('/')[-1] if '/' in subnet else '24'
                
                interface_name = f"vxlan{vni}"
                
                lines.append(f"# {row.get('Nom sous-réseau', '')}")
                lines.append(f"ip link add {interface_name} type vxlan \\")
                lines.append(f"  id {vni} \\")
                lines.append(f"  dev eth0 \\")
                lines.append(f"  group {multicast_ip} \\")
                lines.append(f"  dstport 4789")
                lines.append(f"ip addr add {gateway}/{prefix_len} dev {interface_name}")
                lines.append(f"ip link set {interface_name} up")
                lines.append("")
        
        return "\n".join(lines)
    
    @staticmethod
    def generate_network_plan(data: pd.DataFrame, network_type: str, 
                            max_entries: int = 30) -> str:
        """Génère un plan d'adressage textuel"""
        lines = []
        
        lines.append("=" * 70)
        lines.append(" " * 20 + "PLAN D'ADRESSAGE")
        if len(data) > 0:
            base_net = data.iloc[0].get('Sous-réseau', '').split('/')[0]
            lines.append(" " * 15 + f"Réseau de base: {base_net}")
        lines.append("=" * 70)
        lines.append("")
        
        for idx, row in data.head(max_entries).iterrows():
            nom = row.get('Nom sous-réseau', f'Subnet_{idx}')
            
            if network_type == 'VLAN':
                vlan_id = row.get('VLAN ID', 'N/A')
                vlan_name = row.get('Nom VLAN', 'N/A')
                lines.append(f"{nom} - VLAN {vlan_id} - {vlan_name}")
            else:
                vni = row.get('VNI', 'N/A')
                vxlan_name = row.get('Nom VXLAN', 'N/A')
                lines.append(f"{nom} - VXLAN {vni} - {vxlan_name}")
            
            lines.append(f"  Réseau     : {row.get('Sous-réseau', 'N/A')}")
            lines.append(f"  Plage      : {row.get('Plage IP', 'N/A')}")
            lines.append(f"  Gateway    : {row.get('Gateway', 'N/A')}")
            lines.append(f"  Broadcast  : {row.get('Broadcast', 'N/A')}")
            lines.append(f"  Hôtes      : {row.get('Hosts utilisables', 0)} utilisables")
            lines.append("")
        
        if len(data) > max_entries:
            lines.append(f"... et {len(data) - max_entries} autres sous-réseaux")
        
        return "\n".join(lines)
